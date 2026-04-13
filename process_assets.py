"""
Main Asset Processing Script
Reads Asset.xls and processes investment data
"""

import pandas as pd
from datetime import datetime
from asset_processor import AssetDatabase, AssetAllocator, GainCalculator, TemplateManager
from utils import filter_ticker, get_held_at, empty_to_default, clean_up
import sys
import os
import csv
from dotenv import load_dotenv
import msoffcrypto
import io


class AssetProcessor:
    """Main class to process Asset.xls file"""
    
    def __init__(self, excel_file: str = 'Asset.xls'):
        self.excel_file = excel_file
        # Load environment variables
        load_dotenv()
        self.excel_password = os.getenv('password', '')
        
        # Auto-convert .xls to .xlsx if it's an old format
        if excel_file.endswith('.xls') and not excel_file.endswith('.xlsx'):
            xlsx_file = excel_file + 'x'  # Asset.xls -> Asset.xlsx
            if not os.path.exists(xlsx_file):
                print(f"Converting {excel_file} to modern .xlsx format...")
                self.convert_xls_to_xlsx(excel_file, xlsx_file)
                self.excel_file = xlsx_file
                print(f"Using converted file: {xlsx_file}")
            else:
                print(f"Using existing .xlsx file: {xlsx_file}")
                self.excel_file = xlsx_file
        
        self.db = AssetDatabase(
            host='localhost',
            port=3306,
            user='root',
            password='sa123',
            database='asset'
        )
        self.allocator = AssetAllocator(self.db)
        self.gain_calculator = GainCalculator(self.db)
        self.template_manager = TemplateManager(self.db)

    @staticmethod
    def _parse_currency_value(value):
        """Convert currency-like strings to float, returning None when invalid."""
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)

        text = str(value).strip()
        if text == '' or text == '-':
            return None

        # Handle accounting format like (1,234.56)
        is_negative = text.startswith('(') and text.endswith(')')
        cleaned = text.replace('$', '').replace(',', '').replace('%', '').strip('()').strip()

        try:
            amount = float(cleaned)
            return -amount if is_negative else amount
        except ValueError:
            return None

    def read_trow_csv_entries(self, csv_path: str) -> list:
        """
        Read T. Rowe Price holdings from CSV and return normalized account_ticker entries.

        CSV format contains two sections:
        1) Account holdings table with account type/ticker/market value.
        2) Retirement plan table after a 'TRPRps' marker line where rows are name$amt.
        """
        if not os.path.exists(csv_path):
            raise FileNotFoundError(f"Trow CSV file not found: {csv_path}")

        account_mapping = {
            'Rollover IRA': 'TRPRollover',
            'Individual': 'TRPInv',
            'Roth IRA': 'TRPRoth',
        }

        entries = []
        trprps_by_ticker = {}
        in_rps_section = False

        trprps_mapping = {
            'VANGUARD INST EXT MKT IDX D': 'VIIIX',
            'VANGUARD INST 500 IDX TR D': 'VIEIX',
            'TRP STABLE VALUE COMM TR FD-N': 'Cash',
            'VANGUARD FTSE SOCIAL INDEX I': 'VFTNX',
        }

        with open(csv_path, 'r', newline='', encoding='utf-8-sig') as f:
            reader = csv.reader(f)

            for raw_row in reader:
                row = [cell.strip() if cell is not None else '' for cell in raw_row]
                if not row or all(cell == '' for cell in row):
                    continue

                # Detect second section marker: TRPRps
                if len(row) >= 1 and row[0].strip().lower() == 'trprps':
                    in_rps_section = True
                    continue

                # Backward compatibility with older second-table header.
                if len(row) >= 2 and row[0].lower() == 'investment name' and row[1].lower() == 'amount':
                    in_rps_section = True
                    continue

                if in_rps_section:
                    # Parse rows in the form: Investment Name$123,456.78
                    if len(row) < 1:
                        continue

                    # Amounts include commas and are unquoted in trow.csv, so csv.reader
                    # may split a single logical row into multiple columns.
                    line = ','.join(row).strip()
                    if '$' not in line:
                        continue

                    inv_name, amt_part = line.rsplit('$', 1)
                    inv_name = inv_name.strip()
                    amount = self._parse_currency_value(f'${amt_part.strip()}')
                    if amount is None or amount == 0:
                        continue

                    mapped_ticker = trprps_mapping.get(inv_name)
                    if mapped_ticker:
                        trprps_by_ticker[mapped_ticker] = trprps_by_ticker.get(mapped_ticker, 0.0) + amount
                    elif inv_name.upper().startswith('TRP RETIREMENT BLEND'):
                        trprps_by_ticker['TRRIX'] = trprps_by_ticker.get('TRRIX', 0.0) + amount
                    continue

                # Skip the first section header row.
                if len(row) >= 1 and row[0].lower() == 'account type':
                    continue

                if len(row) < 9:
                    continue

                account_type = row[0]
                ticker = row[2]
                market_value = self._parse_currency_value(row[8])

                account_prefix = account_mapping.get(account_type)
                if not account_prefix:
                    continue

                # Match existing Trow behavior: skip blank ticker and zero/empty value.
                if ticker == '' or market_value is None or market_value == 0:
                    continue

                entries.append({
                    'account_ticker': f"{account_prefix}_{ticker}",
                    'amount': market_value
                })

        for ticker, total_amount in trprps_by_ticker.items():
            if total_amount == 0:
                continue
            entries.append({
                'account_ticker': f'TRPRps_{ticker}',
                'amount': total_amount
            })

        return entries
    
    def read_stocks_csv_entries(self, csv_path: str) -> list:
        """
        Read stock holdings from CSV and calculate Stock = Total - Cash.
        
        CSV format has three columns (no header, tab-delimited):
        1. Account name
        2. Category (Stock, Cash, Total)
        3. Value
        
        Returns a list of entries with account_ticker and amount
        """
        if not os.path.exists(csv_path):
            raise FileNotFoundError(f"Stocks CSV file not found: {csv_path}")
        
        accounts = {}  # Dict to store {account_name: {Stock, Cash, Total}}
        entries = []
        
        with open(csv_path, 'r', newline='', encoding='utf-8-sig') as f:
            reader = csv.reader(f, delimiter='\t')
            
            for raw_row in reader:
                row = [cell.strip() if cell is not None else '' for cell in raw_row]
                
                # Skip empty rows
                if not row or all(cell == '' for cell in row):
                    continue
                
                # Skip rows with insufficient columns
                if len(row) < 3:
                    continue
                
                account_name = row[0].strip()
                category = row[1].strip()
                
                # Skip if account or category is empty
                if not account_name or not category:
                    continue
                
                try:
                    value = float(row[2])
                except (ValueError, IndexError):
                    continue
                
                # Initialize account dict if needed
                if account_name not in accounts:
                    accounts[account_name] = {'Stock': None, 'Cash': None, 'Total': None}
                
                # Store the value by category
                if category in accounts[account_name]:
                    accounts[account_name][category] = value
        
        # Calculate Stock = Total - Cash for each account and create entries
        for account_name, values in accounts.items():
            total = values.get('Total')
            cash = values.get('Cash')
            
            # Skip if we don't have Total and Cash values
            if total is None or cash is None:
                continue
            
            # Add Cash entry if non-zero
            if cash != 0:
                entries.append({
                    'account_ticker': f'{account_name}_Cash',
                    'amount': cash
                })
            
            # Calculate and add stock value
            stock_value = total - cash
            
            # Add Stock entry if non-zero
            if stock_value != 0:
                entries.append({
                    'account_ticker': f'{account_name}_Stock',
                    'amount': stock_value
                })
        
        return entries
    
    def convert_xls_to_xlsx(self, xls_file: str, xlsx_file: str):
        """
        Convert .xls file to .xlsx format
        
        Args:
            xls_file: Source .xls file path
            xlsx_file: Destination .xlsx file path
        """
        try:
            from openpyxl import Workbook
            
            # Decrypt if needed
            if self.excel_password:
                try:
                    decrypted = io.BytesIO()
                    with open(xls_file, 'rb') as f:
                        office_file = msoffcrypto.OfficeFile(f)
                        office_file.load_key(password=self.excel_password)
                        office_file.decrypt(decrypted)
                    decrypted.seek(0)
                    
                    # Read all sheets from .xls
                    xls_data = pd.read_excel(decrypted, sheet_name=None, header=None)
                except:
                    xls_data = pd.read_excel(xls_file, sheet_name=None, header=None)
            else:
                xls_data = pd.read_excel(xls_file, sheet_name=None, header=None)
            
            # Write to .xlsx
            with pd.ExcelWriter(xlsx_file, engine='openpyxl') as writer:
                for sheet_name, df in xls_data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            
            print(f"Successfully converted {xls_file} to {xlsx_file}")
            
        except Exception as e:
            print(f"Error converting file: {e}")
            raise
    
    def update_assetalloc_dates(self, prevdate: str = None, currdate: str = None):
        """Update Assetalloc sheet B1 (prevdate) and C1 (currdate) with direct values"""
        if not prevdate and not currdate:
            return
        
        print(f"\nUpdating Assetalloc sheet dates...")
        try:
            from openpyxl import load_workbook
            import msoffcrypto
            import io
            
            # Load workbook
            wb = load_workbook(self.excel_file, data_only=False)
            
            # Check available sheet names (case-insensitive search)
            sheet_name = None
            for name in wb.sheetnames:
                if name.lower() == 'assetalloc':
                    sheet_name = name
                    break
            
            if not sheet_name:
                print(f"  Warning: Assetalloc sheet not found in {self.excel_file}")
                print(f"  Available sheets: {', '.join(wb.sheetnames)}")
                return
            
            ws = wb[sheet_name]
            # Set direct values (this will replace any formulas)
            if prevdate:
                ws['B1'].value = prevdate
                print(f"  Set B1 (prevdate) = {prevdate}")
            if currdate:
                ws['C1'].value = currdate
                print(f"  Set C1 (currdate) = {currdate}")
            
            wb.save(self.excel_file)
            print(f"  Assetalloc sheet updated successfully")
        except Exception as e:
            print(f"  Error: Could not update Assetalloc dates: {e}")
            import traceback
            traceback.print_exc()
    
    def normalize_full_view(self, sheet_name: str = 'fidfullview', output_file: str = None) -> pd.DataFrame:
        """
        Normalize and aggregate Fidelity data by account
        Reads data from Fidelity.csv instead of Excel sheet
        Converted from VBA normalizefullview() function
        
        Args:
            sheet_name: Deprecated - kept for backward compatibility
            output_file: Optional output Excel file to save results
            
        Returns:
            DataFrame with normalized data
        """
        print(f"Reading Fidelity data from Fidelity.csv...")
        
        # Read Fidelity.csv - it has header row with columns: Account Name, Symbol, Current Value, etc.
        csv_path = os.path.join(os.path.dirname(os.path.abspath(self.excel_file)), 'Fidelity.csv')
        if not os.path.exists(csv_path):
            raise FileNotFoundError(f"Fidelity.csv not found at {csv_path}")
        
        # Read CSV - due to trailing commas in data, columns are shifted by 1
        # We'll use positional indexing to get the correct columns
        df_raw = pd.read_csv(csv_path)
        
        # Account mapping based on Account Name values
        account_mapping = {
            'Individual - TOD': 'FidelityInv',
            'Rollover IRA': 'FidelityIRA',
            'Samir S Doshi - Brokerage Account - 10498558': 'Vanguard',
            'Samir S Doshi - Rollover IRA': 'Vanguard IRA'
        }
        
        account_data = {}  # Dictionary to store aggregated amounts by account
        ticker_data = {}  # Dictionary to aggregate by account_ticker combination
        
        # Process each row using positional indexing
        for idx, row in df_raw.iterrows():
            # Due to trailing commas, columns are shifted:
            # Column 0 (Account Number) contains Account Name
            # Column 1 (Account Name) contains Symbol
            # Column 2 (Symbol) contains Description
            # Column 6 (Last Price Change) contains Current Value
            account_name = row.iloc[0]  # From Account Number column
            symbol = row.iloc[1]  # From Account Name column
            description = row.iloc[2]  # From Symbol column
            value = row.iloc[6]  # From Last Price Change column
            
            # Skip if any required field is missing or empty
            if pd.isna(account_name) or pd.isna(symbol) or pd.isna(value):
                continue
            
            account_name = str(account_name).strip()
            symbol = str(symbol).strip()
            
            if account_name == '' or symbol == '':
                continue
            
            # Handle stock entries where Symbol='Stock'
            if symbol == 'Stock':
                # Keep symbol as 'Stock' so all individual stocks are aggregated
                symbol = 'Stock'
            else:
                # Strip trailing ** (e.g. FCASH** -> FCASH, FDRXX** -> FDRXX)
                symbol = symbol.rstrip('*')
                
                # Map specific symbols to Cash
                if symbol in ['FDRXX', 'FCASH', 'Pending activity', 'VMRXX', 'VMFXX']:
                    symbol = 'Cash'
            
            # Clean up value if it's a string
            if isinstance(value, str):
                value = value.replace('$', '').replace(',', '')
                try:
                    value = float(value)
                except:
                    continue
            
            # Skip zero values
            if value == 0:
                continue
            
            # Map account name to prefix
            account_prefix = account_mapping.get(account_name, None)
            
            # Skip if account not in mapping (handled separately by trow.csv)
            if not account_prefix:
                continue
            
            # Create account_ticker key
            account_ticker = f"{account_prefix}_{symbol}"
            
            # Aggregate by account_ticker (sum duplicates)
            if account_ticker not in ticker_data:
                ticker_data[account_ticker] = 0
            ticker_data[account_ticker] += value
            
            # Aggregate by account
            if account_prefix not in account_data:
                account_data[account_prefix] = 0
            account_data[account_prefix] += value
        
        # Convert ticker_data dictionary to list of results
        results = []
        for account_ticker, amount in ticker_data.items():
            results.append({
                'account_ticker': account_ticker,
                'amount': amount
            })
        
        # Create results DataFrame
        results_df = pd.DataFrame(results)
        
        # Remove entries with 0 value (shouldn't happen but safety check)
        if len(results_df) > 0:
            results_df = results_df[results_df['amount'] != 0]
        
        # Read Trow holdings from CSV and append using the same account_ticker/amount shape.
        try:
            trow_csv_path = os.path.join(os.path.dirname(os.path.abspath(self.excel_file)), 'trow.csv')
            print(f"Reading Trow data from CSV: {trow_csv_path}")
            trow_entries = self.read_trow_csv_entries(trow_csv_path)

            if trow_entries:
                results_df = pd.concat([results_df, pd.DataFrame(trow_entries)], ignore_index=True)
            print(f"Added {len(trow_entries)} entries from trow.csv")
        except Exception as e:
            print(f"Warning: Could not read trow.csv: {e}")
        
        # Read stock account data from CSV and append
        try:
            stocks_csv_path = os.path.join(os.path.dirname(os.path.abspath(self.excel_file)), 'stocks.csv')
            print(f"Reading Stock account data from CSV: {stocks_csv_path}")
            stock_entries = self.read_stocks_csv_entries(stocks_csv_path)
            
            if stock_entries:
                results_df = pd.concat([results_df, pd.DataFrame(stock_entries)], ignore_index=True)
            print(f"Added {len(stock_entries)} entries from stocks.csv")
        except Exception as e:
            print(f"Warning: Could not read stocks.csv: {e}")
        
        # Sort final results
        results_df = results_df.sort_values(by='account_ticker', ascending=True)
        
        # Create summary DataFrame
        summary_df = pd.DataFrame([
            {'account': account, 'total': total}
            for account, total in account_data.items()
        ])
        
        print(f"\nNormalized {len(results_df)} fund entries across {len(summary_df)} accounts")
        print("\nAccount Totals:")
        for _, row in summary_df.iterrows():
            print(f"  {row['account']}: ${row['total']:,.2f}")
        
        # Write results to allaccounts.csv
        try:
            csv_output_path = os.path.join(os.path.dirname(os.path.abspath(self.excel_file)), 'allaccounts.csv')
            results_df.to_csv(csv_output_path, index=False)
            print(f"\nResults written to {csv_output_path}")
        except Exception as e:
            print(f"Warning: Could not write to allaccounts.csv: {e}")
        
        # Save to separate output file if specified
        if output_file:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                results_df.to_excel(writer, sheet_name='Details', index=False)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
            print(f"\nResults also saved to {output_file}")
        
        return results_df, summary_df

    
    def decrypt_and_read_excel(self, sheet_name: str, header=0) -> pd.DataFrame:
        """
        Decrypt password-protected Excel file and read sheet
        
        Args:
            sheet_name: Name of the sheet to read
            header: Row to use as column names (default: 0)
            
        Returns:
            DataFrame with data
        """
        try:
            # Try to read directly first (unprotected file)
            # Use openpyxl with data_only=True to read calculated formula values
            from openpyxl import load_workbook as openpyxl_load
            wb = openpyxl_load(self.excel_file, data_only=True)
            
            # Check if sheet exists
            if sheet_name not in wb.sheetnames:
                available_sheets = ', '.join(wb.sheetnames)
                raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {available_sheets}")
            
            ws = wb[sheet_name]
            
            # Convert to DataFrame manually
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(row)
            
            if header is not None and len(data) > header:
                df = pd.DataFrame(data[header+1:], columns=data[header])
            else:
                df = pd.DataFrame(data)
            
            return df
        except Exception as e:
            # If failed, try with password decryption
            if self.excel_password:
                try:
                    print(f"Attempting to decrypt {self.excel_file}...")
                    decrypted = io.BytesIO()
                    with open(self.excel_file, 'rb') as f:
                        office_file = msoffcrypto.OfficeFile(f)
                        office_file.load_key(password=self.excel_password)
                        office_file.decrypt(decrypted)
                    
                    decrypted.seek(0)
                    
                    # Remove external references from the zip to avoid openpyxl errors
                    from zipfile import ZipFile
                    cleaned = io.BytesIO()
                    with ZipFile(decrypted, 'r') as zin:
                        with ZipFile(cleaned, 'w') as zout:
                            for item in zin.infolist():
                                # Skip external links and related files
                                if 'externalLink' not in item.filename and 'externalReferences' not in item.filename:
                                    data = zin.read(item.filename)
                                    zout.writestr(item, data)
                    
                    cleaned.seek(0)
                    
                    # Use openpyxl to load cleaned file
                    from openpyxl import load_workbook as openpyxl_load
                    wb = openpyxl_load(cleaned, data_only=True)
                    ws = wb[sheet_name]
                    
                    # Convert to DataFrame manually
                    data = []
                    for row in ws.iter_rows(values_only=True):
                        data.append(row)
                    
                    if header is not None and len(data) > header:
                        df = pd.DataFrame(data[header+1:], columns=data[header])
                    else:
                        df = pd.DataFrame(data)
                    
                    print(f"Successfully decrypted and read sheet '{sheet_name}'")
                    return df
                except Exception as decrypt_error:
                    print(f"Error decrypting Excel file: {decrypt_error}")
                    raise
            else:
                print(f"Error reading Excel file: {e}")
                print("No password found in .env file")
                raise
    
    def read_asset_reference_sheet(self, sheet_name: str = 'fullview') -> pd.DataFrame:
        """
        Read asset data from allaccounts.csv
        Columns: account_ticker (e.g., "FidelityInv_FXAIX", "Etrade_Cash"), amount
        
        Args:
            sheet_name: Deprecated parameter (kept for backward compatibility)
            
        Returns:
            DataFrame with columns: Ticker, Amount, HeldAt
        """
        try:
            csv_path = os.path.join(os.path.dirname(os.path.abspath(self.excel_file)), 'allaccounts.csv')
            
            if not os.path.exists(csv_path):
                raise FileNotFoundError(f"allaccounts.csv not found at {csv_path}")
            
            # Read CSV file
            df = pd.read_csv(csv_path)
            
            # Extract data from account_ticker and amount columns
            all_data = []
            
            for idx, row in df.iterrows():
                account_ticker = row.get('account_ticker')
                amount = row.get('amount')
                
                if pd.notna(account_ticker) and pd.notna(amount) and amount != 0:
                    # Clean up amount
                    if isinstance(amount, str):
                        amount = amount.replace('$', '').replace(',', '')
                        try:
                            amount = float(amount)
                        except:
                            continue
                    
                    # Split account_ticker to get account and ticker
                    account_ticker_str = str(account_ticker)
                    if '_' in account_ticker_str:
                        held_at, ticker = account_ticker_str.split('_', 1)
                        all_data.append({
                            'Ticker': ticker,
                            'Amount': amount,
                            'HeldAt': held_at
                        })
            
            result_df = pd.DataFrame(all_data)
            
            print(f"Successfully read {len(result_df)} entries from allaccounts.csv")
            return result_df
        except FileNotFoundError as e:
            print(f"Error: {e}")
            sys.exit(1)
        except Exception as e:
            print(f"Error reading allaccounts.csv: {e}")
            sys.exit(1)
    
    def process_asset_allocation(self, df: pd.DataFrame, as_of_date: datetime, held_at_column: str = 'HeldAt'):
        """
        Process asset allocation from DataFrame
        
        Args:
            df: DataFrame with asset data
            as_of_date: Date for the asset allocation
            held_at_column: Column name containing held at information
        """
        self.db.open_db()
        
        processed_count = 0
        error_count = 0
        error_details = []  # Track detailed error information
        
        # Assuming columns: Ticker (A), Symbol (B), Amount (E), HeldAt (J)
        # Adjust column names based on actual Excel structure
        
        # Track current stock account being processed
        current_stock_account = None
        stock_account_cash = {}  # Track cash amounts for calculating stock value
        stock_accounts = ['Etrade', 'Ameritrade', 'TradeStation', 'Robinhood']
        
        for index, row in df.iterrows():
            try:
                # Get ticker
                ticker = row.get('Ticker', row.get('Symbol', ''))
                if pd.isna(ticker):
                    continue
                
                ticker = str(ticker).strip()
                
                # Check for end marker
                if ticker == "ENDOFPORTFOLIO":
                    break
                
                # Check if this is a stock account header
                if ticker in stock_accounts:
                    current_stock_account = ticker
                    continue
                
                # For stock accounts, process both Cash and Stock rows
                if current_stock_account:
                    if ticker == "Cash":
                        # Process Cash row
                        amount = row.get('Amount', row.get('Value', 0))
                        if pd.isna(amount):
                            amount = 0
                        else:
                            # Clean up amount
                            if isinstance(amount, str):
                                amount = clean_up(amount)
                                amount = float(amount) if amount else 0
                            else:
                                amount = float(amount)
                        
                        # Store cash amount for this account
                        stock_account_cash[current_stock_account] = amount
                        
                        if amount > 0:
                            # Process as Cash
                            held_at = current_stock_account
                            
                            # Get asset ID for Cash
                            query = """
                                SELECT assetid FROM asset 
                                WHERE ticker = %s OR assetname = %s
                            """
                            results = self.db.execute_query(query, ("Cash", "Cash"))
                            
                            if not results:
                                print(f"Warning: Asset not found for ticker Cash")
                                error_count += 1
                            else:
                                asset_id = results[0]['assetid']
                                print(f"Processing: Cash - ${amount:,.2f} at {held_at}")
                                self.allocator.allocate_asset_ref(asset_id, as_of_date, amount, held_at)
                                processed_count += 1
                        continue
                        
                    elif ticker == "Stock":
                        # Skip Stock row (it's a formula, we'll calculate from Total - Cash)
                        continue
                        
                    elif ticker == "Total":
                        # Process Total row - calculate Stock value
                        total_amount = row.get('Amount', row.get('Value', 0))
                        if pd.isna(total_amount):
                            total_amount = 0
                        else:
                            # Clean up amount
                            if isinstance(total_amount, str):
                                total_amount = clean_up(total_amount)
                                total_amount = float(total_amount) if total_amount else 0
                            else:
                                total_amount = float(total_amount)
                        
                        # Calculate Stock = Total - Cash
                        cash_amount = stock_account_cash.get(current_stock_account, 0)
                        stock_amount = total_amount - cash_amount
                        
                        if stock_amount > 0:
                            # Process as Stock
                            held_at = current_stock_account
                            
                            # Get asset ID for Stock
                            query = """
                                SELECT assetid FROM asset 
                                WHERE ticker = %s OR assetname = %s
                            """
                            results = self.db.execute_query(query, ("Stock", "Stock"))
                            
                            if not results:
                                print(f"Warning: Asset not found for ticker Stock")
                                error_count += 1
                                error_details.append(f"Asset not found: Stock (for {held_at})")
                            else:
                                asset_id = results[0]['assetid']
                                print(f"Processing: Stock - ${stock_amount:,.2f} at {held_at}")
                                self.allocator.allocate_asset_ref(asset_id, as_of_date, stock_amount, held_at)
                                processed_count += 1
                        
                        # Reset after Total row
                        current_stock_account = None
                        continue
                    else:
                        # Skip other rows
                        continue
                
                # Skip total/summary rows for regular accounts
                if 'Total' in ticker or 'total' in ticker:
                    continue
                
                # Filter invalid tickers
                ticker = filter_ticker(ticker)
                if not ticker:
                    continue
                
                # Apply fund mapping (e.g., VMRXX -> VMMXX)
                fund_mapping = {
                    'VMRXX': 'VMMXX'
                }
                if ticker in fund_mapping:
                    original_ticker = ticker
                    ticker = fund_mapping[ticker]
                    print(f"Mapped {original_ticker} -> {ticker}")
                
                # Get held at location
                held_at = row.get(held_at_column, row.get('HeldAt', ''))
                if pd.isna(held_at):
                    held_at = ''
                else:
                    held_at = str(held_at).strip()
                
                if not held_at:
                    print(f"Warning: No 'HeldAt' location for ticker {ticker}")
                    continue
                
                # Get amount
                amount = row.get('Amount', row.get('Value', 0))
                if pd.isna(amount):
                    amount = 0
                else:
                    # Clean up amount (remove $ and ,)
                    if isinstance(amount, str):
                        amount = clean_up(amount)
                        amount = float(amount) if amount else 0
                    else:
                        amount = float(amount)
                
                if amount == 0:
                    continue
                
                # Get asset ID from database
                query = """
                    SELECT assetid FROM asset 
                    WHERE ticker = %s OR assetname = %s
                """
                results = self.db.execute_query(query, (ticker, ticker))
                
                if not results:
                    print(f"Warning: Asset not found for ticker {ticker}")
                    error_count += 1
                    error_details.append(f"Asset not found: {ticker}")
                    continue
                
                asset_id = results[0]['assetid']
                
                # Allocate the asset
                print(f"Processing: {ticker} - ${amount:,.2f} at {held_at}")
                self.allocator.allocate_asset_ref(asset_id, as_of_date, amount, held_at)
                processed_count += 1
                
            except Exception as e:
                print(f"Error processing row {index}: {e}")
                error_count += 1
                error_details.append(f"Row {index}: {e}")
                continue
        
        self.db.close_db()
        
        print(f"\n=== Processing Complete ===")
        print(f"Processed: {processed_count} assets")
        print(f"Errors: {error_count}")
        
        if error_count > 0 and error_details:
            print(f"\nError details:")
            for error in error_details:
                print(f"  - {error}")
    
    def compare_dates_report(self, currdate: datetime, datetocompare: datetime, threshold_percent: float = 5.0, show_all: bool = False):
        """
        Compare asset values between two dates and report significant changes
        
        Args:
            currdate: Current date to compare
            datetocompare: Previous date to compare against
            threshold_percent: Minimum percentage change to report (default: 5%)
            show_all: If True, show all changes regardless of threshold (default: False)
        """
        print(f"\n{'='*60}")
        print(f"Significant Changes Report")
        print(f"{'='*60}")
        print(f"Comparing: {datetocompare.strftime('%Y-%m-%d')} vs {currdate.strftime('%Y-%m-%d')}")
        if show_all:
            print(f"Showing: ALL changes\n")
        else:
            print(f"Threshold: {threshold_percent}% change or more\n")
        
        self.db.open_db()
        
        # Query to get asset values by account and ticker for both dates
        query = """
            SELECT 
                ai.asofdate,
                ai.heldat as account,
                a.ticker,
                SUM(aia.amount) as total_amount
            FROM assetinv ai
            JOIN assetinvalloc aia ON ai.assetinvid = aia.assetinvid
            JOIN asset a ON ai.assetid = a.assetid
            WHERE ai.asofdate IN (%s, %s)
            GROUP BY ai.asofdate, ai.heldat, a.ticker
            ORDER BY ai.heldat, a.ticker, ai.asofdate
        """
        
        results = self.db.execute_query(query, 
            (datetocompare.strftime('%Y-%m-%d'), currdate.strftime('%Y-%m-%d')))
        
        # Organize data by account_ticker and also track account totals
        data = {}
        account_totals = {}
        
        for row in results:
            date = row['asofdate']
            account = row['account']
            ticker = row['ticker']
            amount = float(row['total_amount'])
            
            key = f"{account}_{ticker}"
            if key not in data:
                data[key] = {}
            data[key][date] = amount
            
            # Track account totals
            if account not in account_totals:
                account_totals[account] = {}
            if date not in account_totals[account]:
                account_totals[account][date] = 0
            account_totals[account][date] += amount
        
        # Calculate changes
        changes = []
        # Convert to date objects for comparison
        prev_date = datetocompare.date() if hasattr(datetocompare, 'date') else datetocompare
        curr_date = currdate.date() if hasattr(currdate, 'date') else currdate
        
        for key, dates in data.items():
            # dates dict has date/datetime keys from DB - try both formats
            prev_amount = 0
            curr_amount = 0
            for date_key, amount in dates.items():
                # Convert date_key to date for comparison
                dk = date_key.date() if hasattr(date_key, 'date') else date_key
                if dk == prev_date:
                    prev_amount = amount
                if dk == curr_date:
                    curr_amount = amount
            
            account_name = key.split('_')[0]  # Extract account name from key
            
            if prev_amount == 0 and curr_amount > 0:
                # New position
                changes.append({
                    'key': key,
                    'account': account_name,
                    'prev': prev_amount,
                    'curr': curr_amount,
                    'change': curr_amount,
                    'pct_change': 100.0,
                    'type': 'NEW'
                })
            elif prev_amount != 0 and curr_amount == 0:
                # Closed position
                changes.append({
                    'key': key,
                    'account': account_name,
                    'prev': prev_amount,
                    'curr': curr_amount,
                    'change': -prev_amount,
                    'pct_change': -100.0,
                    'type': 'CLOSED'
                })
            elif prev_amount != 0:
                # Changed position (prev may be negative, e.g. margin balance)
                change_amount = curr_amount - prev_amount
                pct_change = (change_amount / abs(prev_amount)) * 100
                
                # Add to list if show_all or if above threshold
                if show_all or abs(pct_change) >= threshold_percent:
                    changes.append({
                        'key': key,
                        'account': account_name,
                        'prev': prev_amount,
                        'curr': curr_amount,
                        'change': change_amount,
                        'pct_change': pct_change,
                        'type': 'CHANGE'
                    })
        
        # Sort by account first, then by absolute percentage change (descending)
        changes.sort(key=lambda x: (x['account'], -abs(x['pct_change'])))
        
        # Calculate account totals and changes
        account_summary = {}
        for account, dates in account_totals.items():
            prev_total = 0
            curr_total = 0
            for date_key, amount in dates.items():
                dk = date_key.date() if hasattr(date_key, 'date') else date_key
                if dk == prev_date:
                    prev_total = amount
                if dk == curr_date:
                    curr_total = amount
            change_total = curr_total - prev_total
            pct_change_total = (change_total / prev_total * 100) if prev_total > 0 else 0
            
            account_summary[account] = {
                'prev': prev_total,
                'curr': curr_total,
                'change': change_total,
                'pct_change': pct_change_total
            }
        
        # Print account summary first
        if account_summary:
            print(f"{'='*60}")
            print("ACCOUNT TOTALS")
            print(f"{'='*60}")
            print(f"{'Account':<30} {'Previous':>12} {'Current':>12} {'Change':>12} {'% Change':>10}")
            print(f"{'-'*30} {'-'*12} {'-'*12} {'-'*12} {'-'*10}")
            
            grand_prev = 0
            grand_curr = 0
            for account in sorted(account_summary.keys()):
                summary = account_summary[account]
                print(f"{account:<30} ${summary['prev']:>10,.2f} ${summary['curr']:>10,.2f} "
                      f"${summary['change']:>10,.2f} {summary['pct_change']:>9.2f}%")
                grand_prev += summary['prev']
                grand_curr += summary['curr']
            grand_change = grand_curr - grand_prev
            grand_pct = (grand_change / grand_prev * 100) if grand_prev > 0 else 0
            print(f"{'-'*30} {'-'*12} {'-'*12} {'-'*12} {'-'*10}")
            print(f"{'TOTAL':<30} ${grand_prev:>10,.2f} ${grand_curr:>10,.2f} "
                  f"${grand_change:>10,.2f} {grand_pct:>9.2f}%")
            print()
        
        # Print individual changes
        if changes:
            print(f"{'='*60}")
            print("INDIVIDUAL HOLDINGS")
            print(f"{'='*60}")
            
            current_account = None
            for i, item in enumerate(changes):
                # Print account header when it changes
                if item['account'] != current_account:
                    # Print summary line for the previous account before switching
                    if current_account is not None:
                        summary = account_summary.get(current_account, {})
                        if summary:
                            print(f"  {'-'*38} {'-'*12} {'-'*12} {'-'*12} {'-'*10}")
                            sprev = summary['prev']
                            scurr = summary['curr']
                            schange = summary['change']
                            spct = summary['pct_change']
                            print(f"  {'TOTAL':<38} ${sprev:>10,.2f} ${scurr:>10,.2f} ${schange:>10,.2f} {spct:>9.2f}%")
                        print()  # Blank line between accounts
                    current_account = item['account']
                    print(f"\n{current_account}:")
                    print(f"{'  Ticker':<38} {'Previous':>12} {'Current':>12} {'Change':>12} {'% Change':>10}")
                    print(f"  {'-'*38} {'-'*12} {'-'*12} {'-'*12} {'-'*10}")
                
                # Extract ticker from key (remove account prefix)
                ticker = '_'.join(item['key'].split('_')[1:])
                prev = item['prev']
                curr = item['curr']
                change = item['change']
                pct = item['pct_change']
                
                if item['type'] == 'NEW':
                    print(f"  {ticker:<38} {'NEW':>12} ${curr:>10,.2f} ${change:>10,.2f} {'NEW':>10}")
                elif item['type'] == 'CLOSED':
                    print(f"  {ticker:<38} ${prev:>10,.2f} {'CLOSED':>12} ${change:>10,.2f} {'CLOSED':>10}")
                else:
                    print(f"  {ticker:<38} ${prev:>10,.2f} ${curr:>10,.2f} ${change:>10,.2f} {pct:>9.2f}%")
                
                # Print summary line after the last item in the last account
                if i == len(changes) - 1 and current_account is not None:
                    summary = account_summary.get(current_account, {})
                    if summary:
                        print(f"  {'-'*38} {'-'*12} {'-'*12} {'-'*12} {'-'*10}")
                        sprev = summary['prev']
                        scurr = summary['curr']
                        schange = summary['change']
                        spct = summary['pct_change']
                        print(f"  {'TOTAL':<38} ${sprev:>10,.2f} ${scurr:>10,.2f} ${schange:>10,.2f} {spct:>9.2f}%")
            
            print(f"\nTotal holdings shown: {len(changes)}")
        else:
            print("No changes found.")
        
        print(f"{'='*60}\n")
        
        self.db.close_db()
    
    def delete_existing_data(self, as_of_date: datetime):
        """
        Delete existing asset data for a given date
        
        Args:
            as_of_date: Date to delete data for
        """
        print(f"Deleting existing data for {as_of_date.strftime('%Y-%m-%d')}...")
        self.allocator.delete_asset_info(as_of_date)
        print("Deletion complete")
    
    def calculate_gains(self, as_of_date: datetime):
        """
        Calculate gains for all assets
        
        Args:
            as_of_date: Date to calculate gains for
        """
        print(f"Calculating gains for {as_of_date.strftime('%Y-%m-%d')}...")
        self.gain_calculator.calculate_gains(as_of_date)
        print("Gain calculation complete")
    
    def refresh_dataconn(self, currdate: datetime = None, datetocompare: datetime = None):
        """
        Refresh dataconn sheet with data from database queries
        
        Args:
            currdate: Current date to update in wkdates table
            datetocompare: Date to compare to update in wkdates table
        """
        print("=" * 60)
        print("Refreshing Dataconn Sheet")
        print("=" * 60)
        
        # Update Assetalloc sheet dates if provided
        if currdate and datetocompare:
            self.update_assetalloc_dates(
                prevdate=datetocompare.strftime('%Y-%m-%d'),
                currdate=currdate.strftime('%Y-%m-%d')
            )
        
        from openpyxl import load_workbook
        
        # Open database connection
        self.db.open_db()
        
        try:
            # Update wkdates table if dates provided
            if currdate or datetocompare:
                print("\nUpdating wkdates table...")
                if currdate:
                    update_query = "UPDATE wkdates SET currdate = %s"
                    self.db.cursor.execute(update_query, (currdate,))
                    print(f"  Updated currdate to {currdate.strftime('%Y-%m-%d')}")
                
                if datetocompare:
                    update_query = "UPDATE wkdates SET datetocompare = %s"
                    self.db.cursor.execute(update_query, (datetocompare,))
                    print(f"  Updated datetocompare to {datetocompare.strftime('%Y-%m-%d')}")
                
                self.db.connection.commit()
            
            # Validate that data exists for both dates
            if currdate and datetocompare:
                print("\nValidating data exists for both dates...")
                
                # Check what dates have data in the views
                date_check_query = """
                    SELECT DISTINCT asofdate FROM totalbyalloctypedate
                    ORDER BY asofdate
                """
                available_dates = self.db.execute_query(date_check_query)
                available_date_strs = set([d['asofdate'].strftime('%Y-%m-%d') if isinstance(d['asofdate'], datetime) else str(d['asofdate']) for d in available_dates])
                
                currdate_str = currdate.strftime('%Y-%m-%d')
                datetocompare_str = datetocompare.strftime('%Y-%m-%d')
                
                print(f"  Available dates in database: {sorted(available_date_strs)}")
                print(f"  Requested currdate: {currdate_str}")
                print(f"  Requested datetocompare: {datetocompare_str}")
                
                missing_dates = []
                if currdate_str not in available_date_strs:
                    missing_dates.append(f"currdate ({currdate_str})")
                if datetocompare_str not in available_date_strs:
                    missing_dates.append(f"datetocompare ({datetocompare_str})")
                
                if missing_dates:
                    error_msg = f"\nERROR: No data found for {' and '.join(missing_dates)}"
                    print(error_msg)
                    print("Please ensure data exists for both dates before running refresh-dataconn.")
                    return
                
                print("  ✓ Data exists for both dates")
            
            # Execute queries
            print("\nExecuting database queries...")
            
            # Query 1: totalbyalloctypedate → A1:C17
            query1 = """
                SELECT totalbyalloctypedate_0.allocdesc, totalbyalloctypedate_0.asofdate, 
                       totalbyalloctypedate_0.`sum(assetinvalloc.amount)`
                FROM totalbyalloctypedate totalbyalloctypedate_0
            """
            results1 = self.db.execute_query(query1)
            print(f"totalbyalloctypedate: {len(results1)} rows")
            
            # Query 2: heldatbydate → F1:H27
            query2 = """
                SELECT heldatbydate_0.heldat, heldatbydate_0.asofdate, 
                       heldatbydate_0.`sum(assetinvalloc.amount)`
                FROM heldatbydate heldatbydate_0
            """
            results2 = self.db.execute_query(query2)
            print(f"heldatbydate: {len(results2)} rows")
            
            # Query 3: cashheldatbydate → J1:L27
            query3 = """
                SELECT cashheldatbydate_0.heldat, cashheldatbydate_0.asofdate, 
                       cashheldatbydate_0.`sum(assetinvalloc.amount)`
                FROM cashheldatbydate cashheldatbydate_0
            """
            results3 = self.db.execute_query(query3)
            print(f"cashheldatbydate: {len(results3)} rows")
            
            # Query 4: assetbydate → O1:U69
            query4 = """
                SELECT * FROM `asset`.`assetbydate` 
                WHERE assetname NOT IN ('Cash', 'Stock')
            """
            results4 = self.db.execute_query(query4)
            print(f"assetbydate: {len(results4)} rows")
            
            # Load Excel workbook
            print(f"\nLoading Excel file: {self.excel_file}")
            wb = load_workbook(self.excel_file)
            
            # Get DataConn sheet (note: capital D and C)
            if 'DataConn' not in wb.sheetnames:
                ws = wb.create_sheet('DataConn')
            else:
                ws = wb['DataConn']
            
            # Clear existing data in all the columns where we write
            print("Clearing existing data in DataConn sheet...")
            # Clear columns A-C (rows 1-100), F-H (rows 1-100), J-L (rows 1-100), O-U (rows 1-100)
            for row in range(1, 101):
                # Clear A1:C100
                ws.cell(row, 1).value = None
                ws.cell(row, 2).value = None
                ws.cell(row, 3).value = None
                # Clear F1:H100
                ws.cell(row, 6).value = None
                ws.cell(row, 7).value = None
                ws.cell(row, 8).value = None
                # Clear J1:L100
                ws.cell(row, 10).value = None
                ws.cell(row, 11).value = None
                ws.cell(row, 12).value = None
                # Clear O1:U100
                for col in range(15, 22):  # O=15 to U=21
                    ws.cell(row, col).value = None
            
            print("Writing data to DataConn sheet...")
            
            # Write Query 1 results to A1:C17
            if results1:
                for row_idx, row_data in enumerate(results1, start=1):
                    ws.cell(row_idx, 1, row_data.get('allocdesc'))
                    ws.cell(row_idx, 2, row_data.get('asofdate'))
                    ws.cell(row_idx, 3, row_data.get('sum(assetinvalloc.amount)'))
            
            # Write Query 2 results to F1:H27
            if results2:
                for row_idx, row_data in enumerate(results2, start=1):
                    ws.cell(row_idx, 6, row_data.get('heldat'))
                    ws.cell(row_idx, 7, row_data.get('asofdate'))
                    ws.cell(row_idx, 8, row_data.get('sum(assetinvalloc.amount)'))
            
            # Write Query 3 results to J1:L27
            if results3:
                for row_idx, row_data in enumerate(results3, start=1):
                    ws.cell(row_idx, 10, row_data.get('heldat'))
                    ws.cell(row_idx, 11, row_data.get('asofdate'))
                    ws.cell(row_idx, 12, row_data.get('sum(assetinvalloc.amount)'))
            
            # Write Query 4 results to O1:U69
            if results4:
                for row_idx, row_data in enumerate(results4, start=1):
                    # Get all column names dynamically
                    col_offset = 15  # Column O = 15
                    for col_idx, (key, value) in enumerate(row_data.items()):
                        ws.cell(row_idx, col_offset + col_idx, value)
            
            # Save workbook
            wb.save(self.excel_file)
            print(f"\nDataConn sheet updated successfully in {self.excel_file}")
            
        finally:
            self.db.close_db()
        
        print("=" * 60)
        print("Refresh Complete!")
        print("=" * 60)
    
    def fix_external_references(self):
        """
        Fix external workbook references in formulas
        Note: This only fixes formula references. Excel may still show a warning
        about external links, but the formulas will work correctly.
        """
        print("=" * 60)
        print("Fixing External Workbook References")
        print("=" * 60)
        
        from openpyxl import load_workbook
        import re
        import os
        import shutil
        
        # Create backup first
        backup_file = self.excel_file + '.backup'
        if os.path.exists(self.excel_file):
            shutil.copy2(self.excel_file, backup_file)
            print(f"Created backup: {backup_file}")
        
        # Load Excel workbook with data_only=False to preserve formulas
        print(f"\nLoading Excel file: {self.excel_file}")
        wb = load_workbook(self.excel_file, data_only=False, keep_vba=True)
        
        # Pattern to match external workbook references
        pattern_numbered = r'\[\d+\]'  # [1], [2], etc.
        pattern_fin = r"'\[fin\.xlsx\][^']*'!"
        pattern_fin_simple = r'\[fin\.xlsx\][^!]*!'
        
        fixed_count = 0
        fin_count = 0
        
        # Check all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_fixes = 0
            sheet_fin_fixes = 0
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        old_value = cell.value
                        new_value = old_value
                        
                        # Remove numbered external workbook references like [1], [2], etc.
                        if re.search(r'\[\d+\]', old_value):
                            new_value = re.sub(pattern_numbered, '', new_value)
                            sheet_fixes += 1
                            fixed_count += 1
                        
                        # Remove fin.xlsx references
                        if 'fin.xlsx' in old_value.lower():
                            new_value = re.sub(pattern_fin, '#REF!', new_value, flags=re.IGNORECASE)
                            new_value = re.sub(pattern_fin_simple, '#REF!', new_value, flags=re.IGNORECASE)
                            sheet_fin_fixes += 1
                            fin_count += 1
                        
                        if new_value != old_value:
                            cell.value = new_value
            
            if sheet_fixes > 0:
                print(f"  {sheet_name}: Fixed {sheet_fixes} numbered references")
            if sheet_fin_fixes > 0:
                print(f"  {sheet_name}: Removed {sheet_fin_fixes} fin.xlsx references")
        
        # Save workbook
        print("\nSaving workbook...")
        wb.save(self.excel_file)
        wb.close()
        
        print(f"\nFixed {fixed_count} numbered references and {fin_count} fin.xlsx references")
        print("\nNote: Excel may still show an external links warning on open.")
        print("You can safely update or break those links in Excel's Data > Edit Links menu.")
        print("=" * 60)
        print("Fix Complete!")
        print("=" * 60)
    
    def run_full_process(self, as_of_date: datetime, sheet_name: str = 'fullview', 
                        delete_existing: bool = True, calculate_gains: bool = False):
        """
        Run the full asset processing workflow
        
        Args:
            as_of_date: Date for the asset allocation
            sheet_name: Name of the sheet to read
            delete_existing: Whether to delete existing data first
            calculate_gains: Whether to calculate gains after allocation (default: False)
        """
        print("=" * 60)
        print("Asset Processing Workflow")
        print("=" * 60)
        print(f"Excel File: {self.excel_file}")
        print(f"Sheet: {sheet_name}")
        print(f"As of Date: {as_of_date.strftime('%Y-%m-%d')}")
        print("=" * 60)
        
        # Step 1: Delete existing data if requested
        if delete_existing:
            self.delete_existing_data(as_of_date)
        
        # Step 2: Read Excel data
        print("\nReading Excel file...")
        df = self.read_asset_reference_sheet(sheet_name)
        print(f"Read {len(df)} rows from Excel")
        
        # Step 3: Process asset allocation
        print("\nProcessing asset allocation...")
        self.process_asset_allocation(df, as_of_date)
        
        # Step 4: Calculate gains if requested
        if calculate_gains:
            self.calculate_gains(as_of_date)
        
        print("\n" + "=" * 60)
        print("Processing Complete!")
        print("=" * 60)

    def show_unique_dates(self, after_date: datetime = None):
        """
        Show unique dates for which there is data, optionally filtered after a given date
        
        Args:
            after_date: Optional date to filter results - only show dates after this date
        """
        print(f"\n{'='*60}")
        print(f"Available Data Dates")
        print(f"{'='*60}")
        if after_date:
            print(f"Showing dates after: {after_date.strftime('%Y-%m-%d')}\n")
        else:
            print(f"Showing all available dates\n")
        
        self.db.open_db()
        
        # Query to get unique dates from assetinv table
        if after_date:
            query = """
                SELECT DISTINCT asofdate 
                FROM assetinv 
                WHERE asofdate > %s
                ORDER BY asofdate DESC
            """
            results = self.db.execute_query(query, (after_date.strftime('%Y-%m-%d'),))
        else:
            query = """
                SELECT DISTINCT asofdate 
                FROM assetinv 
                ORDER BY asofdate DESC
            """
            results = self.db.execute_query(query)
        
        self.db.close_db()
        
        if not results:
            if after_date:
                print(f"No data found after {after_date.strftime('%Y-%m-%d')}")
            else:
                print("No data found in database")
            return
        
        # Display the dates
        print(f"Found {len(results)} unique date(s):\n")
        for i, row in enumerate(results, 1):
            date_val = row['asofdate']
            # Handle both datetime and date objects
            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                date_str = str(date_val)
            print(f"  {i:2d}. {date_str}")
        
        print(f"\n{'='*60}")


def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Process Asset.xls file and update database')
    parser.add_argument('--file', '-f', default='Asset.xls', 
                       help='Path to Excel file (default: Asset.xls)')
    parser.add_argument('--sheet', '-s', default='fullview',
                       help='Sheet name to process (default: fullview)')
    parser.add_argument('--date', '-d', 
                       help='As-of date in YYYY-MM-DD format (default: today)')
    parser.add_argument('--no-delete', action='store_true',
                       help='Do not delete existing data before processing')
    parser.add_argument('--with-gains', action='store_true',
                       help='Calculate gains after allocation (not default)')
    parser.add_argument('--gains-only', action='store_true',
                       help='Only calculate gains, skip allocation')
    parser.add_argument('--delete-only', action='store_true',
                       help='Only delete data, skip allocation and gains')
    parser.add_argument('--process', action='store_true',
                       help='Run main asset allocation workflow (default if no other mode specified)')
    parser.add_argument('--normalize', action='store_true',
                       help='Normalize full view data and aggregate by account')
    parser.add_argument('--normalize-sheet', default='fidfullview',
                       help='Sheet name for normalize operation (default: fidfullview)')
    parser.add_argument('--output', '-o',
                       help='Output Excel file for normalize results')
    parser.add_argument('--refresh-dataconn', action='store_true',
                       help='Refresh dataconn sheet with database query results')
    parser.add_argument('--compare-dates', action='store_true',
                       help='Compare asset values between two dates and report significant changes')
    parser.add_argument('--show-all', action='store_true',
                       help='Show all changes (ignore threshold) when using --compare-dates')
    parser.add_argument('--threshold', type=float, default=5.0,
                       help='Percentage threshold for significant changes (default: 5.0)')
    parser.add_argument('--currdate',
                       help='Current date for wkdates table and assetref N4 (YYYY-MM-DD)')
    parser.add_argument('--datetocompare',
                       help='Date to compare for wkdates table and assetref N3 (YYYY-MM-DD)')
    parser.add_argument('--fix-references', action='store_true',
                       help='Fix external workbook references in formulas')
    parser.add_argument('--show-dates', action='store_true',
                       help='Show unique dates for which there is data in the database')
    parser.add_argument('--after-date',
                       help='Filter dates to show only those after this date (YYYY-MM-DD). Use with --show-dates')
    
    args = parser.parse_args()
    
    # Parse date
    if args.date:
        try:
            as_of_date = datetime.strptime(args.date, '%Y-%m-%d')
        except ValueError:
            print("Error: Invalid date format. Use YYYY-MM-DD")
            sys.exit(1)
    else:
        as_of_date = datetime.now()
    
    # Parse currdate and datetocompare
    currdate = None
    datetocompare = None
    if args.currdate:
        try:
            currdate = datetime.strptime(args.currdate, '%Y-%m-%d')
        except ValueError:
            print("Error: Invalid currdate format. Use YYYY-MM-DD")
            sys.exit(1)
    
    if args.datetocompare:
        try:
            datetocompare = datetime.strptime(args.datetocompare, '%Y-%m-%d')
        except ValueError:
            print("Error: Invalid datetocompare format. Use YYYY-MM-DD")
            sys.exit(1)
    
    # Parse after_date if provided
    after_date = None
    if args.after_date:
        try:
            after_date = datetime.strptime(args.after_date, '%Y-%m-%d')
        except ValueError:
            print("Error: Invalid after-date format. Use YYYY-MM-DD")
            sys.exit(1)
    
    # Create processor
    processor = AssetProcessor(args.file)
    
    # Execute based on flags
    if args.show_dates:
        # Show unique dates with optional filtering
        processor.show_unique_dates(after_date=after_date)
    elif args.normalize:
        # Update Assetalloc dates
        processor.update_assetalloc_dates(
            prevdate=datetocompare.strftime('%Y-%m-%d') if datetocompare else None,
            currdate=as_of_date.strftime('%Y-%m-%d') if as_of_date else None
        )
        # Run normalize full view
        processor.normalize_full_view(
            sheet_name=args.normalize_sheet,
            output_file=args.output
        )
    elif args.refresh_dataconn:
        # Refresh dataconn sheet
        processor.refresh_dataconn(currdate=currdate, datetocompare=datetocompare)
    elif args.compare_dates:
        # Compare dates and report significant changes
        if not currdate or not datetocompare:
            print("Error: --compare-dates requires --currdate and --datetocompare")
            sys.exit(1)
        processor.compare_dates_report(currdate=currdate, datetocompare=datetocompare, 
                                      threshold_percent=args.threshold, show_all=args.show_all)
    elif args.fix_references:
        # Fix external workbook references
        processor.fix_external_references()
    elif args.delete_only:
        processor.delete_existing_data(as_of_date)
    elif args.gains_only:
        processor.calculate_gains(as_of_date)
    elif args.process or not any([args.normalize, args.refresh_dataconn, 
                                   args.compare_dates, args.fix_references, args.delete_only, 
                                   args.gains_only, args.show_dates]):
        # Update Assetalloc dates
        processor.update_assetalloc_dates(
            prevdate=datetocompare.strftime('%Y-%m-%d') if datetocompare else None,
            currdate=as_of_date.strftime('%Y-%m-%d') if as_of_date else None
        )
        # Run main allocation workflow (default)
        processor.run_full_process(
            as_of_date=as_of_date,
            sheet_name=args.sheet,
            delete_existing=not args.no_delete,
            calculate_gains=args.with_gains
        )


if __name__ == '__main__':
    main()
